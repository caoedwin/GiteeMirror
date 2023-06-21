from django.shortcuts import render,redirect
from django.views.decorators.csrf import csrf_exempt
import datetime,os

from .forms import DriverList
from .forms import ToolList
from .models import DriverList_M,ToolList_M
from app01.models import ProjectinfoinDCT, UserInfo
# from .models import DriverList_M
# from .models import ToolList_M
from django.http import HttpResponse
import datetime,json,simplejson,requests,time
from requests_ntlm import HttpNtlmAuth

# Create your views here.
@csrf_exempt
def DriverList_upload(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
        weizhi="XQM/DriverList_upload"
    DriverList_M_lists = DriverList(request.POST)
    result='00'
    Existlist=[{'Customer': 'Customer', 'Project': 'Project','Phase0': 'Phase', 'Name': 'Driver/Utility/Firmware/Application Name',
                     'Function': 'Function', 'Vendor': 'Vendor', 'Version': 'Version', 'Bios': 'Bios', 'Image': 'Image', 'Driver': 'Driver'}]
    DriverList_M_dic={}
    result=4

    if request.method == "POST":
        if 'type' in request.POST:
            err_ok = 0
            xlsxlist = request.POST.get('upload')
            n = 0

            for i in simplejson.loads(xlsxlist):
                n += 1
                # print(i)
                if 'Customer' in i.keys():
                    Customer = i['Customer']
                else:
                    Customer = ''
                if 'Project' in i.keys():
                    Project = i['Project']
                else:
                    Project = ''
                if 'Phase' in i.keys():
                    Phase0 = i['Phase']
                else:
                    Phase0 = ''
                if 'Driver/Utility/Firmware/Application Name' in i.keys():
                    Name = i['Driver/Utility/Firmware/Application Name']
                else:
                    Name = ''
                if 'Function' in i.keys():
                    Function = i['Function']
                else:
                    Function = ''
                if 'Vendor' in i.keys():
                    Vendor=i['Vendor']
                else:
                    Vendor = ''
                if 'Version' in i.keys():
                    Version = i['Version']
                else:
                    Version = ''
                if 'BIOS' in i.keys():
                    Bios = i['BIOS']
                else:
                    Bios = ''
                if 'Image' in i.keys():
                    Image = i['Image']
                else:
                    Image = ''
                if 'Driver' in i.keys():
                    Driver = i['Driver']
                else:
                    Driver = ''
                # print(len(Version))
                # print(Driver)
                check_dic = {'Customer': Customer, 'Project':Project,'Phase0':Phase0, 'Name': Name,
                     'Function': Function, 'Vendor': Vendor, 'Version':Version, 'BIOS': Bios, 'Image': Image,'Driver': Driver}
                check_list = DriverList_M.objects.filter(**check_dic)
                if check_list:
                   err_ok=1
                   DriverList_M_dic['Customer']=Customer
                   DriverList_M_dic['Project'] = Project
                   DriverList_M_dic['Phase0'] = Phase0
                   DriverList_M_dic['Name'] = Name
                   DriverList_M_dic['Function'] = Function
                   DriverList_M_dic['Vendor'] = Vendor
                   DriverList_M_dic['Version'] = Version
                   DriverList_M_dic['BIOS'] = Bios
                   DriverList_M_dic['Image'] = Image
                   DriverList_M_dic['Driver'] = Driver
                   Existlist.append(DriverList_M_dic)
                   continue
                else:
                    DriverList_Mmodule = DriverList_M()
                    DriverList_Mmodule.Customer = Customer
                    DriverList_Mmodule.Project = Project
                    DriverList_Mmodule.Phase0 = Phase0
                    DriverList_Mmodule.Name = Name
                    DriverList_Mmodule.Function = Function
                    DriverList_Mmodule.Vendor = Vendor
                    DriverList_Mmodule.Version = Version
                    DriverList_Mmodule.BIOS = Bios
                    DriverList_Mmodule.Image = Image
                    DriverList_Mmodule.Driver = Driver
                    DriverList_Mmodule.editor = request.session.get('user_name')
                    DriverList_Mmodule.edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    DriverList_Mmodule.save()
            datajason = {
            'err_ok': err_ok,
            'content': Existlist
            }
            # print(datajason)
            return HttpResponse(json.dumps(datajason), content_type="application/json")
        if 'Upload' in request.POST:
            if DriverList_M_lists.is_valid():
                Customer = DriverList_M_lists.cleaned_data['Customer']
                Project = DriverList_M_lists.cleaned_data['Project']
                Phase0 = DriverList_M_lists.cleaned_data['Phase0']
                Name = DriverList_M_lists.cleaned_data['Name']
                Function = DriverList_M_lists.cleaned_data['Function']
                Vendor = DriverList_M_lists.cleaned_data['Vendor']
                Version = DriverList_M_lists.cleaned_data['Version']
                Bios = DriverList_M_lists.cleaned_data['Bios']
                Image = DriverList_M_lists.cleaned_data['Image']
                Driver = DriverList_M_lists.cleaned_data['Driver']

                check_dic = {'Customer': Customer, 'Project': Project, 'Phase0': Phase0,'Name': Name, 'Function': Function,
                            'Vendor': Vendor, 'Version': Version, 'BIOS': Bios, 'Image': Image, 'Driver': Driver}
                if DriverList_M.objects.filter(**check_dic):
                    result = 1
                else:
                    DriverList_Mmodule = DriverList_M()
                    DriverList_Mmodule.Customer = Customer
                    DriverList_Mmodule.Project = Project
                    DriverList_Mmodule.Phase0 = Phase0
                    DriverList_Mmodule.Name = Name
                    DriverList_Mmodule.Function = Function
                    DriverList_Mmodule.Vendor = Vendor
                    DriverList_Mmodule.Version = Version
                    DriverList_Mmodule.BIOS = Bios
                    DriverList_Mmodule.Image = Image
                    DriverList_Mmodule.Driver = Driver
                    DriverList_Mmodule.editor = request.session.get('user_name')
                    DriverList_Mmodule.edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    DriverList_Mmodule.save()
                    result = 0
            else:
                cleanData = DriverList_M_lists.errors
        DriverList_M_lists = DriverList()
        return render(request, 'DriverTool/DriverList_upload.html', locals())
    return render(request, 'DriverTool/DriverList_upload.html',locals())

import pandas as pd
import pprint
from pathlib import Path
import os, sys
def read_excel(src_file,header=0,sheetnum=0):
    df = pd.read_excel(src_file, header=header, sheet_name=int(sheetnum)).iloc[:,
         0:]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
    # # 显示所有列
    pd.set_option('display.max_columns', None)
    # # 显示所有行
    # pd.set_options('display.max_rows', None)
    # pprint.pprint(df)
    dataexcel = df.values[:, :]
    # datatest = []
    # for i in dataexcel:
    #     ls = []
    #     for j in i:
    #         ls.append(j)
    #     datatest.append(ls)
    # print(list(df.columns))
    # pprint.pprint(datatest)
    df = df.fillna('')  # 替换 Nan, 否则没有双引号的Nan，json.dumps(data)时虽然不报错，但是传到前端反序列化后无法获取数据
    excel_dic = df.to_dict('records')
    print("111", excel_dic)
    hangnum = 1
    for i in excel_dic:
        i["dataid"] = hangnum
        hangnum += 1
    key_data = list(df.columns)
    # pprint.pprint(excel_dic)
    # df.to_excel('C:/media/ABOTestPlan/upload.xlsx', sheet_name="sheet1", index=False,
    #             engine='openpyxl')
    # with pd.ExcelWriter(src_file, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
    #     df.to_excel(writer, sheet_name='Sheet1', index=False)  # engine="openpyxl"
    from openpyxl import load_workbook
    #读取所有批注
    workbook = load_workbook(src_file)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    comments = []
    rownum = 0
    for row in worksheet.rows:
        cellnum = 0
        for cell in row:
            if cell.comment:
                comments.append([rownum, cellnum, cell.comment.text])
            cellnum += 1
        rownum += 1
    print(comments)
    return excel_dic,key_data
from openpyxl import load_workbook
def save_exel(save_data,src_file,header=0,sheetnum=0):
    df1 = pd.read_excel(src_file, sheet_name=None)
    sheetname = list(df1)
    df2 = pd.DataFrame(save_data)
    print("2222", df2)
    foo = pd.DataFrame({
        'temp': ['message1', 'message2', 'message3'],
        'var2': [1, 2, 3],
        'col3': [4, 5, 6]
    })

    # Setup a DataFrame with corresponding hover values
    # tooltips_df = pd.DataFrame({
    #     'temp': ['i am message 1', 'i am foo', 'i am lala'],
    #     'var2': ' another random message',
    #     'col3': 'more random messages'
    # })
    #
    # # Assign tooltips
    # foo.style.set_tooltips(tooltips_df)
    # foo.to_excel('C:/media/ABOTestPlan/tips.xlsx', sheet_name="sheet1", index=False)
    # print("333", foo)
    # excel_writer = pd.ExcelWriter(r'C:\Users\Administrator\Desktop\test2.xlsx')  # 定义writer，选择文件（文件可以不存在,相当于新建文件)
    # data1.to_excel(excel_writer, sheet_name='sheet_data1')
    # data2.to_excel(excel_writer, sheet_name='sheet_data2')
    # excel_writer.save()  # 保存文件   ---data1和data2都在，原数据被覆盖
    #无法保证打结果的始终在最左列
    # with pd.ExcelWriter(src_file, mode='a', engine='openpyxl') as writer:
    #     wb = writer.book  # openpyxl.workbook.workbook.Workbook 获取所有sheet
    #     wb.remove(wb[sheetname[0]])  # 删除需要覆盖的sheet
    #     # print(wb.sheetnames)
    #     # df2.to_excel(writer, sheet_name=sheetname[0], index=True)  ##sheet st3的内容更新成st1值
    # with pd.ExcelWriter(src_file, mode='a', engine='openpyxl') as writer:
    #     df2.to_excel(writer, sheet_name=sheetname[0], index=False)  ##sheet st3的内容更新成st1值

    #无法保存公式，样式，注解
    excel_list = [df2]
    for i in sheetname:
        if i != sheetname[0]:
            excel_list.append(pd.read_excel(src_file, sheet_name=i))
    with pd.ExcelWriter(src_file, engine='openpyxl') as writer:
        num = 0
        for i in excel_list:
            i.to_excel(writer, sheet_name=sheetname[num], index=False)  ##sheet st3的内容更新成st1值
            num += 1









@csrf_exempt
def DriverList_edit(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi="XQM/DriverList_edit"
    # status='0'
    mock_data = [
        # {'id': 1, "Name": "Intel Serial I/O driver", "Function": "I/O", "Vendor": "Intel","Version": "Inbox","Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 2, "Name": "Intel WinPE Serial I/O driver", "Function": "I/O", "Vendor": "Intel","Version": "30.100.1727.1", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 3, "Name": "Intel Chipset driver", "Function": "Chipset", "Vendor": "Intel","Version": "10.1.17809.8096(DM:10.1.15.5)", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 4, "Name": "Intel Rapid Storage Technology driver", "Function": "IRST", "Vendor": "Intel","Version": "17.0.0.1072", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 5, "Name": "Intel WinRE Rapid Storage TechnologyO driver", "Function": "IRST", "Vendor": "Intel","Version": "17.0.0.1072", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 6, "Name": "Intel Management Engine Interface driver", "Function": "MEI", "Vendor": "Intel","Version": "1851.12.0.1193(CSME:12.0.23.1311)", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 7, "Name": "Intel Dynamic Platform Thermal driver", "Function": "DPTF", "Vendor": "Intel","Version": "8.5.10103.7263", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"},
        #          {'id': 8, "Name": "Intel Dynamic Platform Thermal Framework driver", "Function": "ISST", "Vendor": "Intel","Version": "10.23.0.2241", "Project": "EL451_S540-14IWL", "Driver": "V1.11", "Image": "v25 GML"}
    ]
    selectItem = {
        # "C38(NB)": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "C38(AIO)": [{"Project": "EL535", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL536", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL537", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL538", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "A39": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "Other": [{"Project": "ELMV2", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV3", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV4", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}]
    }

    dr = [
        # {"Driver": "v0.8"}, {"Driver": "v1.0"}, {"Driver": "v1.11"}
    ]
    image = [
        # {"Image": "v24.1 GMl"}, {"Image": "v22.0 GMl"}
    ]
    excel_dic = []
    key_list = []
    canExport = 1
    canEdit = 1

    for i in DriverList_M.objects.all().values('Customer').distinct().order_by('Customer'):
        Customerlist=[]
        # print(type(i))
        for j in DriverList_M.objects.filter(Customer=i['Customer']).values('Project').distinct().order_by('Project'):
            Projectlist = {}

            dic={'Customer':i['Customer'],'Project':j['Project']}
            Phaselist=[]
            for l in DriverList_M.objects.filter(**dic).values('Phase0').distinct().order_by('Phase0'):
                Phaselist.append(l['Phase0'])
            Projectlist['Project']=j["Project"]
            Projectlist['Phase0']=Phaselist
            Customerlist.append(Projectlist)
        selectItem[i['Customer']]=Customerlist


    if request.method == "POST":

        # print(request.POST, request.method)
        # print(request.body)
        # print(type(request.body),type(request.POST))
        # responseData = json.loads(request.body)
        # print(responseData)
        # if request.POST.get('isGetData') == 'first':

        # test = request.POST
        # for i in test:
        #     print(test[i])
        if request.POST.get('isGetData')=='PHASE':
            dic_phase={'Customer':request.POST.get('Customer'),'Project':request.POST.get('Project'),'Phase0':request.POST.get('Phase')}
            # print(dic_phase)
            for i in DriverList_M.objects.filter(**dic_phase).values('Driver').distinct().order_by('Driver'):
                # print(i)
                dr.append({'Driver' : i['Driver']})
            for i in DriverList_M.objects.filter(**dic_phase).values('Image').distinct().order_by('Image'):
                # print(i)
                image.append({'Image' : i['Image']})
        if request.POST.get('isGetData') == 'SEARCH':
            src_file = "C:/media/ABOTestPlan/do.xlsx"
            excel_dic = read_excel(src_file)[0]
            # print(type(excel_dic))
            key_list = read_excel(src_file)[1]
            save_exel(excel_dic,src_file)

            Customer= request.POST.get('Customer')
            Project=request.POST.get('Project')
            Phase=request.POST.get('Phase')
            Driver=request.POST.get('Driver')
            Image=request.POST.get('Image')
            dic={}
            if Customer:
                dic['Customer']=Customer
            if Project:
                dic['Project']=Project
            if Phase:
                dic['Phase0']=Phase
            if Driver:
                dic['Driver']=Driver
                # print(Driver,len(Driver))
            if Image:
                dic['Image']=Image
            for i in DriverList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor":i.Vendor,
                                  "Version": i.Version,"Project":i.Project, "Bios": i.BIOS, "Driver": i.Driver, "Image": i.Image,'Customer':i.Customer,'Phase':i.Phase0,})

        if request.POST.get('isGetData') == 'SAVE':
            # print(request.POST)
            id=request.POST.get('rows[id]')
            DriverList_Mmodule=DriverList_M.objects.get(id=id)
            DriverList_Mmodule.Customer=request.POST.get('rows[Customer]')
            DriverList_Mmodule.Project = request.POST.get('rows[Project]')
            DriverList_Mmodule.Phase0 = request.POST.get('rows[Phase]')
            DriverList_Mmodule.Name = request.POST.get('rows[Name]')
            DriverList_Mmodule.Function = request.POST.get('rows[Function0]')
            DriverList_Mmodule.Vendor = request.POST.get('rows[Vendor]')
            DriverList_Mmodule.Version = request.POST.get('rows[Version]')
            DriverList_Mmodule.BIOS = request.POST.get('rows[Bios]')
            DriverList_Mmodule.Image = request.POST.get('rows[Image]')
            DriverList_Mmodule.Driver = request.POST.get('rows[Driver]')
            DriverList_Mmodule.editor = request.session.get('user_name')
            DriverList_Mmodule.edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            DriverList_Mmodule.save()

            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            Driver = request.POST.get('Driver')
            Image = request.POST.get('Image')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            if Driver:
                dic['Driver'] = Driver
            if Image:
                dic['Image'] = Image
            for i in DriverList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor": i.Vendor,
                                  "Version": i.Version, "Project": i.Project, "Bios": i.BIOS, "Driver": i.Driver, "Image": i.Image,
                                  'Customer': i.Customer, 'Phase': i.Phase0, })
        if request.POST.get('isGetData') == 'DELETE':
            # print(request.POST)
            id = request.POST.get('id')
            DriverList_Mmodule = DriverList_M.objects.get(id=id)
            DriverList_Mmodule.delete()
            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            Driver = request.POST.get('Driver')
            Image = request.POST.get('Image')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            if Driver:
                dic['Driver'] = Driver
            if Image:
                dic['Image'] = Image
            for i in DriverList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor": i.Vendor,
                                  "Version": i.Version, "Project": i.Project, "Bios": i.BIOS, "Driver": i.Driver, "Image": i.Image,
                                  'Customer': i.Customer, 'Phase': i.Phase0, })
        if 'MUTICANCEL' in str(request.body):
            responseData = json.loads(request.body)
            # print(responseData)
            for i in responseData['params']:
                DriverList_M.objects.get(id=i).delete()
            Customer = responseData['Customer']
            Project = responseData['Project']
            Phase = responseData['Phase']
            Driver = responseData['Driver']
            Image = responseData['Image']
            # print(Customer,Project)
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            if Driver:
                dic['Driver'] = Driver
            if Image:
                dic['Image'] = Image
            # print(dic)
            for i in DriverList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor": i.Vendor,
                                  "Version": i.Version, "Project": i.Project, "Bios": i.BIOS, "Driver": i.Driver, "Image": i.Image,
                                  'Customer': i.Customer, 'Phase': i.Phase0, })
            # status='1'
        data = {
            "err_ok": "0",
            "content": mock_data,
            "select": selectItem,
            "selectedDriver": dr,
            "selectedImage": image,
            "excel_dic": excel_dic,
            "key_list": key_list,
            "canExport": canExport,
            "canEdit": canEdit,
            # "status":status
        }
        # print(type(json.dumps(data)),json.dumps(data))
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'DriverTool/DriverList_edit.html', locals())


@csrf_exempt
def DriverList_search(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi="XQM/DriverList_search"

    mock_data = [
        # {"Name": "Intel Serial I/O driver", "Function": "I/O", "Vendor": "Intel", "Version": "Inbox","Project": "EL451_S540-14IWL", "Driver": "v1.11", "Image": "v25 GM1"},
        #          {"Name": "Intel WinPE Serial I/O driver", "Function": "I/O", "Vendor": "Intel","Version": "30.100.1727.1", "Project": "EL451_S540-14IWL", "Driver": "v1.11", "Image": "v25 GM1"},
        #          {"Name": "Intel Chipset driver", "Function": "Chipset", "Vendor": "Intel","Version": "10.1.17809.8096(DM:10.1.15.5)", "Project": "EL451_S540-14IWL", "Driver": "v1.11","Image": "v25 GM1"},
        #          {"Name": "Intel Rapid Storage Technology driver", "Function": "IRST", "Vendor": "Intel","Version": "17.0.0.1072", "Project": "EL451_S540-14IWL", "Driver": "v1.11", "Image": "v25 GM1"},
        #          {"Name": "Intel WinRE Rapid Storage Technology driver", "Function": "IRST", "Vendor": "Intel","Version": "17.0.0.1072", "Project": "EL451_S540-14IWL", "Driver": "v1.11", "Image": "v25 GM1"},
        #          {"Name": "Intel Management Engine Interface driver", "Function": "MEI", "Vendor": "Intel","Version": "1851.12.0.1193", "Project": "EL451_S540-14IWL", "Driver": "v1.11","Image": "v25 GM1"},
        #          {"Name": "Intel Dynamic Platform Thermal Framework driver", "Function": "DPTF", "Vendor": "Intel","Version": "8.5.10103.7263", "Project": "EL451_S540-14IWL", "Driver": "v1.11","Image": "v25 GM1"}
    ]

    selectItem = {
        # "C38(NB)":[{"Project":"EL531", "Phase0":["B(FVT)","C(SIT)","INV"]},{"Project":"EL532","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL533","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL534","Phase0":["B(FVT)","C(SIT)","INV"]}],
        #           "C38(AIO)":[{"Project":"EL535", "Phase0":["B(FVT)","C(SIT)","INV"]},{"Project":"EL536","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL537","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL538","Phase0":["B(FVT)","C(SIT)","INV"]}],
        #           "A39": [{"Project":"EL531", "Phase0":["B(FVT)","C(SIT)","INV"]},{"Project":"EL532","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL533","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"EL534","Phase0":["B(FVT)","C(SIT)","INV"]}],
        #           "Other":[{"Project":"ELMV2", "Phase0":["B(FVT)","C(SIT)","INV"]},{"Project":"ELMV3","Phase0":["B(FVT)","C(SIT)","INV"]}, {"Project":"ELMV4","Phase0":["B(FVT)","C(SIT)","INV"]}]
    }
    sear = []
    #dr=[{"Driver":"v0.8"},{"Driver":"v1.0"},{"Driver":"v1.11"}],
    dr=[
        # {"Driver":"v0.8"},{"Driver":"v1.0"},{"Driver":"v1.11"}
    ]
    image=[
        # {"Image":"v24.1 GMl"},{"Image":"v22.0 GMl"}
    ]
    for i in DriverList_M.objects.all().values('Customer').distinct().order_by('Customer'):
        Customerlist=[]
        # print(type(i))
        for j in DriverList_M.objects.filter(Customer=i['Customer']).values('Project').distinct().order_by('Project'):
            Projectlist = {}

            dic={'Customer':i['Customer'],'Project':j['Project']}
            Phaselist=[]
            for l in DriverList_M.objects.filter(**dic).values('Phase0').distinct().order_by('Phase0'):
                Phaselist.append(l['Phase0'])
            Projectlist['Project']=j["Project"]
            Projectlist['Phase0']=Phaselist
            Customerlist.append(Projectlist)
        selectItem[i['Customer']]=Customerlist
    canExport = 0
    roles = []
    onlineuser = request.session.get('account')
    # print(UserInfo.objects.get(account=onlineuser))
    for i in UserInfo.objects.get(account=onlineuser).role.all():
        roles.append(i.name)
    for i in roles:
        if 'admin' in i:
            # editPpriority = 4
            canExport = 1
        # elif 'DQA' in i and 'edit' in i:
        #     canExport = 1
    if request.method == "POST":
        # if request.POST.get("isGetData")=="PHASE":
        #     dic_phase = {'Customer': request.POST.get('Customer'), 'Project': request.POST.get('Project'),
        #                  'Phase0': request.POST.get('Phase')}
        #     # print(dic_phase)
        #     for i in DriverList_M.objects.filter(**dic_phase).values('Driver').distinct().order_by('Driver'):
        #         # print(i)
        #         dr.append({'Driver': i['Driver']})
        #     for i in DriverList_M.objects.filter(**dic_phase).values('Image').distinct().order_by('Image'):
        #         # print(i)
        #         image.append({'Image': i['Image']})
        if request.POST.get('isGetData') == 'SEARCHALERT':
            Customer = request.POST.get('Customer')
            Prolist = []
            if Customer:
                for i in DriverList_M.objects.filter(Customer=Customer).values("Project").distinct().order_by("Project"):
                    Prolist.append(i["Project"])
            else:
                for i in DriverList_M.objects.all().values("Project").distinct().order_by("Project"):
                    Prolist.append(i["Project"])
            for i in Prolist:
                if ProjectinfoinDCT.objects.filter(ComPrjCode=i).first():
                    sear.append({
                        "YEAR": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Year, "COMPRJCODE": i,
                        "PrjEngCode1": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PrjEngCode1,
                        "PrjEngCode2": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PrjEngCode2,
                        "PROJECT": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().ProjectName,
                        "SIZE": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Size,
                        "CPU": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().CPU,
                        "PLATFORM": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Platform,
                        "VGA": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().VGA,
                        "OS SUPPORT": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().OSSupport,
                        "Type": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Type,
                        "PPA": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PPA,
                        "PQE": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PQE,
                        "SS": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().SS,
                        "LD": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().LD,
                        "DQA PL": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().DQAPL,
                    })
                else:
                    sear.append({
                        "YEAR": "", "COMPRJCODE": i,
                        "CUSPRJCODE": "",
                        "ProjectName": "",
                        "SIZE": "",
                        "CPU": "",
                        "PLATFORM": "",
                        "VGA": "",
                        "OS SUPPORT": "",
                        "Type": "",
                        "PPA": "",
                        "PQE": "",
                        "SS": "",
                        "LD": "",
                        "DQA PL": "",
                    })
            # print(sear)
        if request.POST.get('isGetData') == 'clickdetail':
            print(request.POST)
            Customer = request.POST.get('Customer')
            Project = request.POST.get('row.COMPRJCODE')
            Phase = request.POST.get('Phase')
            Driver = request.POST.get('Driver')
            Image = request.POST.get('Image')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            if Driver:
                dic['Driver'] = Driver
            if Image:
                dic['Image'] = Image
            for i in DriverList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor": i.Vendor,
                                  "Version": i.Version, "Project": i.Project, "Bios": i.BIOS, "Driver": i.Driver, "Image": i.Image,
                                  'Customer': i.Customer, 'Phase': i.Phase0, })
        # if request.POST.get('isGetData') == 'SEARCH':
        #     Customer = request.POST.get('Customer')
        #     Project = request.POST.get('Project')
        #     Phase = request.POST.get('Phase')
        #     Driver = request.POST.get('Driver')
        #     Image = request.POST.get('Image')
        #     dic = {}
        #     if Customer:
        #         dic['Customer'] = Customer
        #     if Project:
        #         dic['Project'] = Project
        #     if Phase:
        #         dic['Phase0'] = Phase
        #     if Driver:
        #         dic['Driver'] = Driver
        #     if Image:
        #         dic['Image'] = Image
        #     for i in DriverList_M.objects.filter(**dic):
        #         mock_data.append({'id': i.id, "Name": i.Name, "Function0": i.Function, "Vendor": i.Vendor,
        #                           "Version": i.Version, "Project": i.Project, "Driver": i.Driver, "Image": i.Image,
        #                           'Customer': i.Customer, 'Phase': i.Phase0, })


        data = {
            "err_ok": "0",
            "content": mock_data,
            "select": selectItem,
            #"select0":selectList,
            "sear": sear,
            "selectedDriver":dr,
            "selectedImage":image,
            'canExport': canExport,
        }
        return HttpResponse(json.dumps(data), content_type="application/json")

    return render(request, 'DriverTool/DriverList_search.html', locals())

@csrf_exempt
def ToolList_upload(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi = "XQM/ToolList_upload"
    ToolList_upload = ToolList(request.POST)
    ToolList_M_lists = [{'Customer': 'Customer', 'Project': 'Project',
                     'Phase0': 'Phase', 'Vendor': 'Vendor', 'Version': 'Version',
                     'ToolName': 'ToolName', 'TestCase': 'TestCase'}]
    result = '00'
    ToolList_M_dic = {}
    result = 4
    if request.method == "POST":
        if 'type' in request.POST:
            err_ok = 0
            xlsxlist = request.POST.get('upload')
            n = 0

            for i in simplejson.loads(xlsxlist):
                n += 1
                if 'Customer' in i.keys():
                    Customer=i['Customer']
                else:
                    Customer=''
                if 'Project' in i.keys():
                    Project=i['Project']
                else:
                    Project=''
                if 'Phase' in i.keys():
                    Phase0=i['Phase']
                else:
                    Phase0=''
                if 'Vendor' in i.keys():
                    Vendor=i['Vendor']
                else:
                    Vendor=''
                if 'Version' in i.keys():
                    Version=i['Version']
                else:
                    Version=''
                if 'ToolName' in i.keys():
                    ToolName = i['ToolName']
                else:
                    ToolName = ''
                if 'TestCase' in i.keys():
                    TestCase = i['TestCase']
                else:
                    TestCase = ''
                # print(len(Version))
                check_dic = {'Customer': Customer, 'Project': Project,
                             'Phase0': Phase0, 'Vendor': Vendor,
                             'Version': Version, 'ToolName': ToolName, 'TestCase': TestCase}
                check_list = ToolList_M.objects.filter(**check_dic)
                if check_list:
                    err_ok = 1
                    ToolList_M_dic['Customer'] = Customer
                    ToolList_M_dic['Project'] = Project
                    ToolList_M_dic['Phase0'] = Phase0
                    ToolList_M_dic['Vendor'] = Vendor
                    ToolList_M_dic['Version'] = Version
                    ToolList_M_dic['ToolName'] = ToolName
                    ToolList_M_dic['TestCase'] = TestCase
                    ToolList_M_lists.append(ToolList_M_dic)
                    continue
                else:
                    ToolList_Mmodule = ToolList_M()
                    ToolList_Mmodule.Customer = Customer
                    ToolList_Mmodule.Project = Project
                    ToolList_Mmodule.Phase0 = Phase0
                    ToolList_Mmodule.Vendor = Vendor
                    ToolList_Mmodule.Version = Version
                    ToolList_Mmodule.ToolName = ToolName
                    ToolList_Mmodule.TestCase = TestCase
                    ToolList_Mmodule.editor = request.session.get('user_name')
                    ToolList_Mmodule.edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ToolList_Mmodule.save()
            datajason = {
                'err_ok': err_ok,
                'content': ToolList_M_lists
            }
            return HttpResponse(json.dumps(datajason), content_type="application/json")
        if 'Upload' in request.POST:
            if ToolList_upload.is_valid():  # 必须要先验证否则提示object错误没有attribute 'cleaned_data'
                Customer = ToolList_upload.cleaned_data['Customer']
                Project = ToolList_upload.cleaned_data['Project']
                Phase0 = ToolList_upload.cleaned_data['Phase0']
                Vendor = ToolList_upload.cleaned_data['Vendor']
                Version = ToolList_upload.cleaned_data['Version']
                ToolName = ToolList_upload.cleaned_data['ToolName']
                TestCase = ToolList_upload.cleaned_data['TestCase']
                check_dic={'Customer':Customer,'Project':Project,'Phase0':Phase0,'Vendor':Vendor,
                           'Version':Version,'ToolName':ToolName,'TestCase':TestCase}
                if ToolList_M.objects.filter(**check_dic):
                    result=1
                else:
                    ToolList_Mmodule = ToolList_M()
                    ToolList_Mmodule.Customer = Customer
                    ToolList_Mmodule.Project = Project
                    ToolList_Mmodule.Phase0 = Phase0
                    ToolList_Mmodule.Vendor = Vendor
                    ToolList_Mmodule.Version = Version
                    ToolList_Mmodule.ToolName = ToolName
                    ToolList_Mmodule.TestCase = TestCase
                    ToolList_Mmodule.editor = request.session.get('user_name')
                    ToolList_Mmodule.edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ToolList_Mmodule.save()
                    message_CDM = "Upload Successfully"
                    result = 0
                return render(request, 'DriverTool/ToolList_upload.html',
                                  {'weizhi': weizhi, 'Skin': Skin, 'ToolList_upload':ToolList(),
                                   'result': result})
            else:
                cleanData = ToolList_upload.errors
        return render(request, 'DriverTool/ToolList_upload.html',
                      {'weizhi': weizhi, 'Skin': Skin, 'ToolList_upload': ToolList(), 'result': result})

    return render(request, 'DriverTool/ToolList_upload.html', locals())

@csrf_exempt
def ToolList_edit(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi = "Reliability Test Data/ToolList_edit"
    mock_data=[
        # {'id':1,"ToolName":"STPM","TestCase":"Stress","Vendor":"Compal initernal","Project":"EL451_S540-14IWL","Version":"V2.5.0.0"},
        #        {'id':2,"ToolName": "3DMark11", "TestCase": "Stress", "Vendor": "FutureMark", "Project": "EL451_S540-14IWL", "Version": "v1.0.132"},
        #        {'id':3,"ToolName": "3DMark", "TestCase": "Stress Performance", "Vendor": "FutureMark", "Project": "EL451_S540-14IWL", "Version": "v2.8.6546"},
        #        {'id':4,"ToolName": "Bench loop", "TestCase": "Stress", "Vendor": "", "Project": "EL451_S540-14IWL", "Version": "v1.2.6"},
        #        {'id':5,"ToolName": "BurnInTest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL", "Version": "V8.1.1022.0"},
        #        {'id':6,"ToolName": "STPM_ThermalCherk(C38)", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL", "Version": "v1.6.1.0"},
        #        {'id':7,"ToolName": "Pwrtest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL", "Version": "v10586_TH2"},
        #        {'id':8,"ToolName": "PingTest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL", "Version": "v3.6.7.11"},
        #        {'id':9,"ToolName": "EC Tool", "TestCase": "N/A", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v08.00e"}
    ]
    selectItem = {
        # "C38(NB)": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "C38(AIO)": [{"Project": "EL535", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL536", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL537", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL538", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "A39": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "Other": [{"Project": "ELMV2", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV3", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV4", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}]
    }
    for i in ToolList_M.objects.all().values('Customer').distinct().order_by('Customer'):
        Customerlist=[]
        # print(type(i))
        for j in ToolList_M.objects.filter(Customer=i['Customer']).values('Project').distinct().order_by('Project'):
            Projectlist = {}

            dic={'Customer':i['Customer'],'Project':j['Project']}
            Phaselist=[]
            for l in ToolList_M.objects.filter(**dic).values('Phase0').distinct().order_by('Phase0'):
                Phaselist.append(l['Phase0'])
            Projectlist['Project']=j["Project"]
            Projectlist['Phase0']=Phaselist
            Customerlist.append(Projectlist)
        selectItem[i['Customer']]=Customerlist
    if request.method == "POST":
        # print (request.POST,request.method)
        if request.POST.get('isGetData')=='SEARCH':
            Customer=request.POST.get('Customer')
            Project=request.POST.get('Project')
            Phase=request.POST.get('Phase')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            for i in ToolList_M.objects.filter(**dic):
                mock_data.append({'id':i.id,'ToolName':i.ToolName,'TestCase':i.TestCase,'Vendor':i.Vendor,
                                  'Project':i.Project,'Version':i.Version,'Customer':i.Customer,'Phase':i.Phase0})
        if request.POST.get('isGetData')=='SAVE':
            # print(request.POST)
            id=request.POST.get('rows[id]')
            editdata=ToolList_M.objects.get(id=id)
            # editdata.Customer=request.POST.get('row[Customer]')
            editdata.Project = request.POST.get('rows[Project]')
            editdata.Phase0 = request.POST.get('rows[Phase]')
            editdata.ToolName = request.POST.get('rows[ToolName]')
            editdata.TestCase = request.POST.get('rows[TestCase]')
            editdata.Vendor = request.POST.get('rows[Vendor]')
            editdata.Version = request.POST.get('rows[Version]')
            editdata.editor = request.session.get('user_name')
            editdata.edit_time = datetime.datetime.now().strftime("%Y-%m-d %H:%M:%S")
            editdata.save()

            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            for i in ToolList_M.objects.filter(**dic):
                mock_data.append({'id':i.id,'ToolName': i.ToolName, 'TestCase': i.TestCase, 'Vendor': i.Vendor,
                                  'Project': i.Project, 'Version': i.Version, 'Customer': i.Customer,
                                  'Phase': i.Phase0})
        if request.POST.get('isGetData')=="DELETE":
            # print(request.POST)
            # print(request.POST.get('id'))
            ToolList_M.objects.get(id=request.POST.get('id')).delete()

            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            for i in ToolList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, 'ToolName': i.ToolName, 'TestCase': i.TestCase, 'Vendor': i.Vendor,
                                  'Project': i.Project, 'Version': i.Version, 'Customer': i.Customer,
                                  'Phase': i.Phase0})
        if 'MUTICANCEL' in str(request.body):
            responseData = json.loads(request.body)
            # print(responseData)
            for i in responseData['params']:
                ToolList_M.objects.get(id=i).delete()
            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            for i in ToolList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, 'ToolName': i.ToolName, 'TestCase': i.TestCase, 'Vendor': i.Vendor,
                                  'Project': i.Project, 'Version': i.Version, 'Customer': i.Customer,
                                  'Phase': i.Phase0})

        data = {
            "err_ok": "0",
            "content":mock_data,
            "select":selectItem
        }
        # print(data)
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'DriverTool/ToolList_edit.html', locals())

@csrf_exempt
def ToolList_search(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi = "XQM/ToolList_search"
    mock_data = [
        # {"ToolName": "STPM", "TestCase": "Stress", "Vendor": "Compal initernal", "Project": "EL451_S540-14IWL","Version": "V2.5.0.0"},
        #          {"ToolName": "3DMark11", "TestCase": "Stress", "Vendor": "FutureMark", "Project": "EL451_S540-14IWL","Version": "v1.0.132"},
        #          {"ToolName": "3DMark", "TestCase": "Stress Performance", "Vendor": "FutureMark","Project": "EL451_S540-14IWL", "Version": "v2.8.6546"},
        #          {"ToolName": "Bench loop", "TestCase": "Stress", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v1.2.6"},
        #          {"ToolName": "BurnInTest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "V8.1.1022.0"},
        #          {"ToolName": "STPM_ThermalCherk(C38)", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v1.6.1.0"},
        #          {"ToolName": "Pwrtest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v10586_TH2"},
        #          {"ToolName": "PingTest", "TestCase": "", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v3.6.7.11"},
        #          {"ToolName": "EC Tool", "TestCase": "N/A", "Vendor": "", "Project": "EL451_S540-14IWL","Version": "v08.00e"}
    ]
    sear = []
    selectItem={
        # "C38(NB)":["EL531", "EL532", "EL533", "EL534"],"C38(AIO)":["FL535", "FL536", "FL537", "FL538"],"A39":["FL531", "FL532", "FL533", "FL534"],"Other":["ELMV2", "ELMV3", "ELMV4"]
                }
    for i in ToolList_M.objects.all().values('Customer').distinct().order_by('Customer'):
        Customerlist=[]
        # print(type(i))
        for j in ToolList_M.objects.filter(Customer=i['Customer']).values('Project').distinct().order_by('Project'):
            Projectlist = {}

            dic={'Customer':i['Customer'],'Project':j['Project']}
            Phaselist=[]
            for l in ToolList_M.objects.filter(**dic).values('Phase0').distinct().order_by('Phase0'):
                Phaselist.append(l['Phase0'])
            Projectlist['Project']=j["Project"]
            Projectlist['Phase0']=Phaselist
            Customerlist.append(Projectlist)
        selectItem[i['Customer']]=Customerlist
    canExport = 0
    roles = []
    onlineuser = request.session.get('account')
    # print(UserInfo.objects.get(account=onlineuser))
    for i in UserInfo.objects.get(account=onlineuser).role.all():
        roles.append(i.name)
    for i in roles:
        if 'admin' in i:
            # editPpriority = 4
            canExport = 1
        elif 'DQA' in i:
            canExport = 1
    if request.method == "POST":
        if request.POST.get('isGetData') == 'SEARCHALERT':
            Customer = request.POST.get('Customer')
            Prolist = []
            if Customer:
                for i in ToolList_M.objects.filter(Customer=Customer).values("Project").distinct().order_by("Project"):
                    Prolist.append(i["Project"])
            else:
                for i in ToolList_M.objects.all().values("Project").distinct().order_by("Project"):
                    Prolist.append(i["Project"])
            for i in Prolist:
                if ProjectinfoinDCT.objects.filter(ComPrjCode=i).first():
                    sear.append({
                        "YEAR": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Year, "COMPRJCODE": i,
                        "PrjEngCode1": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PrjEngCode1,
                        "PrjEngCode2": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PrjEngCode2,
                        "PROJECT": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().ProjectName,
                        "SIZE": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Size,
                        "CPU": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().CPU,
                        "PLATFORM": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Platform,
                        "VGA": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().VGA,
                        "OS SUPPORT": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().OSSupport,
                        "Type": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().Type,
                        "PPA": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PPA,
                        "PQE": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().PQE,
                        "SS": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().SS,
                        "LD": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().LD,
                        "DQA PL": ProjectinfoinDCT.objects.filter(ComPrjCode=i).first().DQAPL,
                    })
                else:
                    sear.append({
                        "YEAR": "", "COMPRJCODE": i,
                        "CUSPRJCODE": "",
                        "ProjectName": "",
                        "SIZE": "",
                        "CPU": "",
                        "PLATFORM": "",
                        "VGA": "",
                        "OS SUPPORT": "",
                        "Type": "",
                        "PPA": "",
                        "PQE": "",
                        "SS": "",
                        "LD": "",
                        "DQA PL": "",
                    })
            # print(sear)
        if request.POST.get('isGetData') == 'clickdetail':
            Customer = request.POST.get('Customer')
            Project = request.POST.get('row.COMPRJCODE')
            Phase = request.POST.get('Phase')
            dic = {}
            if Customer:
                dic['Customer'] = Customer
            if Project:
                dic['Project'] = Project
            if Phase:
                dic['Phase0'] = Phase
            for i in ToolList_M.objects.filter(**dic):
                mock_data.append({'id': i.id, 'ToolName': i.ToolName, 'TestCase': i.TestCase, 'Vendor': i.Vendor,
                                  'Project': i.Project, 'Version': i.Version, 'Customer': i.Customer,
                                  'Phase': i.Phase0})
        # if request.POST.get('isGetData')=='SEARCH':
        #     Customer=request.POST.get('Customer')
        #     Project=request.POST.get('Project')
        #     Phase=request.POST.get('Phase')
        #     dic = {}
        #     if Customer:
        #         dic['Customer'] = Customer
        #     if Project:
        #         dic['Project'] = Project
        #     if Phase:
        #         dic['Phase0'] = Phase
        #     for i in ToolList_M.objects.filter(**dic):
        #         mock_data.append({'id':i.id,'ToolName':i.ToolName,'TestCase':i.TestCase,'Vendor':i.Vendor,
        #                           'Project':i.Project,'Version':i.Version,'Customer':i.Customer,'Phase':i.Phase0})
        data = {
            "err_ok": "0",
            "content":mock_data,
            "select":selectItem,
            "sear": sear,
            'canExport': canExport,
        }
        # print(data)
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'DriverTool/ToolList_search.html', locals())