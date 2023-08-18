from django.shortcuts import render,redirect,reverse
from django.views.decorators.csrf import csrf_exempt
import datetime,os

from app01.models import ProjectinfoinDCT, UserInfo
# from .models import DriverList_M
# from .models import ToolList_M
from django.http import HttpResponse
import datetime,json,simplejson,requests,time
import pandas as pd
import pprint
from pathlib import Path
import os, sys, shutil
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as ExcelColumn
from openpyxl.utils import column_index_from_string as Col2Int
from openpyxl.comments import Comment
from openpyxl.styles import Side, Border, Font, Alignment, PatternFill

import threading
from concurrent.futures import ThreadPoolExecutor,ProcessPoolExecutor
#ThreadPoolExecutor线程池ProcessPoolExecutor进程池


import numpy as np
# Create your views here.

from win32com.client import Dispatch
import pythoncom
def just_open(filename):
    print("just open")
    pythoncom.CoInitialize()
    xlApp = Dispatch("Excel.Application")

    xlApp.Visible = False

    xlBook = xlApp.Workbooks.Open(filename)

    xlBook.Save()

    xlBook.Close()

def Copy_forders(forderpath, folder_path_Sys):
    print("Copy_forders")
    #判断是否有人在搜索
    # cycles = 0
    # while cycles < 10:
    #     if os.path.exists(folder_path_Sys):
    #         time.sleep(1)
    #         cycles += 1
    #     else:
    #         break
    # 经常用到的(如果文件夹不存在，则创建该文件夹)
    if not os.path.exists(folder_path_Sys):
        os.makedirs(folder_path_Sys)
    if os.path.exists(folder_path_Sys):
        # root 所指的是当前正在遍历的这个文件夹的本身的地址
        # dirs 是一个 list，内容是该文件夹中所有的目录的名字(不包括子目录)
        # files 同样是 list, 内容是该文件夹中所有的文件(不包括子目录)
        for root, dirs, files in os.walk(forderpath):
            for file in files:
                # print(file)
                if not file.startswith('~$'):
                    src_file = os.path.join(root, file)
                    shutil.copy(src_file, folder_path_Sys)
                    # print(src_file)
    return


# def Del_forders(folder_path_Sys):
#     cycles = 0
#     while cycles < 10:
#         try:
#             shutil.rmtree(folder_path_Sys)
#         except Exception as e:
#             print(e)
#             time.sleep(1)
#             cycles += 1
#         else:
#             break
#         return


def read_excel(src_file,header=0,sheetnum=1):
    """
        注意：如果该工作簿是用openpyxl创建的，并且在创建后未曾用Microsoft
        Excel打开过，那么想要读取公式计算结果是无法得到正确结果的，只会读出None。

        其原因如下：

        　　当xlsx表格被生成并在Excel程序中打开并保存之后（这个过程Excel会把公式结果计算出来），该文件中所有单元格附带有两套值，一套是公式全都没有计算的，一套是公式计算了结果的。此时，openpyxl以data_only = False打开可以读取公式，以data_only = True打开可以得到公式计算出的结果。

        　　如果openpyxl创建的工作簿没有被Excel打开并保存，则只有data_only = False的一套值，没有公式计算结果的那一套值。所以data_only = True读取会得到None。

        　　另外：如果用openpyxl的data_only = True状态打开文件，并且最后用save()
        函数保存了后，则xlsx文件中的公式会被替换为计算结果（即消除了公式）。

        　　而如果用openpyxl的data_only = False状态下打开文件，最后用save()
        函数保存了的话，原xlsx文件也会只剩下data_only = False的那套值（即公式），另一套（data_only = True）的值会丢失，如想重新获得两套值，则仍旧需要手动用Excel程序打开该文件并保存一次。

        　　那么能否不用手动用Excel程序打开就能读取公式计算结果呢？可以的！使用win32com自动打开文件并保存一下就好了。代码如下：
    """
    # just_open(src_file)
    print("read_excel")
    # work_book = openpyxl.load_workbook('test.xlsx', data_only=True)
    workbook = load_workbook(src_file, data_only=False)
    first_sheet = workbook.get_sheet_names()[1]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    excel_fx = []
    rownum = 0
    for row in worksheet.rows:
        cellnum = 0
        for cell in row:
            cell_value = worksheet.cell(row=rownum + 1, column=cellnum + 1).value
            # print(cell_value)
            if cell_value:
                if str(cell_value).startswith('='):
                    # print(cell_value)
                    excel_fx.append([rownum + 1, cellnum + 1, cell_value])
            cellnum += 1
        rownum += 1
    workbook2 = load_workbook(src_file, data_only=True)
    first_sheet2 = workbook2.get_sheet_names()[1]
    worksheet2 = workbook2.get_sheet_by_name(first_sheet2)
    # print(excel_fx,1)
    if excel_fx:
        cell_value = worksheet2.cell(row=excel_fx[0][0], column=excel_fx[0][1]).value
        # print(cell_value, 2)
        if not cell_value:
            # print("open")
            just_open(src_file)
    #打开just_open后，pandas也能督导公式运算后的值
    df = pd.read_excel(src_file, header=header, sheet_name=int(sheetnum),keep_default_na=False).iloc[:,
         0:]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
    # # 显示所有列
    pd.set_option('display.max_columns', None)
    # # 显示所有行
    pd.set_option('display.max_rows', None)
    # pprint.pprint(df)
    dataexcel = df.values[:, :]
    # datatest = []
    # for i in dataexcel:
    #     ls = []
    #     for j in i:
    #         ls.append(j)
    #     datatest.append(ls)
    # print(list(df.columns))
    # pprint.pprint(datatest)
    df = df.fillna('')  # 替换 Nan, 否则没有双引号的Nan，json.dumps(data)时虽然不报错，但是传到前端反序列化后无法获取数据
    excel_dic = df.to_dict('records')
    # print("111", excel_dic)
    hangnum = 1
    for i in excel_dic:
        i["dataid"] = hangnum
        hangnum += 1
    key_data = list(df.columns)
    # pprint.pprint(excel_dic)
    # df.to_excel('C:/media/ABOTestPlan/upload.xlsx', sheet_name="sheet1", index=False,
    #             engine='openpyxl')
    # with pd.ExcelWriter(src_file, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
    #     df.to_excel(writer, sheet_name='Sheet1', index=False)  # engine="openpyxl"

    #读取所有批注


    comments = []

    rownum = 0
    for row in worksheet.rows:
        cellnum = 0
        for cell in row:
            if cell.comment:
                # print(row, cell)
                comments.append([rownum-1, cellnum, cell.comment.text, key_data[cellnum]])
            cellnum += 1
        rownum += 1
    # print(comments)
    for i in excel_dic:
        i["comments"] = []
        i["commentsedit"] = {}
    for i in comments:
        # print(i,excel_dic[i[0]])
        excel_dic_num = i[0]
        if "comments" in excel_dic[excel_dic_num].keys():
            excel_dic[excel_dic_num]["comments"].append(i)#for summary,需要将所有comments心是在一起
            excel_dic[excel_dic_num]["commentsedit"][i[3]] = i#for edit，编辑修改内容，需要将修改的注释内容精确到单元格

    return excel_dic,key_data, comments

def info_excel(src_file,header=0,sheetnum0=0,sheetnum=1):
    # print("info_excel")
    # print(datetime.datetime.now())
    df = pd.read_excel(src_file, header=header, sheet_name=int(sheetnum0)).iloc[:,
         :]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
    # # 显示所有列
    pd.set_option('display.max_columns', None)
    # # 显示所有行
    # pd.set_options('display.max_rows', None)
    # pprint.pprint(df)
    dataexcel = df.values[:, :]
    # datatest = []
    # for i in dataexcel:
    #     ls = []
    #     for j in i:
    #         ls.append(j)
    #     datatest.append(ls)
    # print(list(df.columns))
    # pprint.pprint(datatest)
    df = df.fillna('')  # 替换 Nan, 否则没有双引号的Nan，json.dumps(data)时虽然不报错，但是传到前端反序列化后无法获取数据
    excel_dic = df.to_dict('records')
    # print("111", excel_dic)
    hangnum = 1
    for i in excel_dic:
        i["dataid"] = hangnum
        hangnum += 1
    key_data = list(df.columns)
    # print(datetime.datetime.now(), 2)

    # 您可以使用参数keep_default_na 和na_values 手动设置所有NA 值docs：防止pandas在读取excel时删除'NA‘字符串
    df = pd.read_excel(src_file, header=header, sheet_name=int(sheetnum), keep_default_na=False).iloc[42:,
         1:]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
    # # 显示所有列
    pd.set_option('display.max_columns', None)
    # # 显示所有行
    pd.set_option('display.max_rows', None)
    # df = df.fillna('?')  # 替换 Nan, 否则没有双引号的Nan，json.dumps(data)时虽然不报错，但是传到前端反序列化后无法获取数据, None,NA,NAN,NAT,Null都被认为是缺失值
    # df = df.fillna(method='ffill')
    key_data = list(df.columns)
    df = df.replace("", "?")
    # pprint.pprint(df)

    P_value = df.eq('P').sum()
    All_tongjidata_P = pd.DataFrame([P_value.values], columns=P_value.index).to_dict('records')
    F_value = df.eq('F').sum()
    All_tongjidata_F = pd.DataFrame([F_value.values], columns=F_value.index).to_dict('records')
    B_value = df.eq('B').sum()
    All_tongjidata_B = pd.DataFrame([B_value.values], columns=B_value.index).to_dict('records')
    NS_value = df.eq('NS').sum()
    All_tongjidata_NS = pd.DataFrame([NS_value.values], columns=NS_value.index).to_dict('records')
    # NaN_value = df.eq('?').sum()
    # All_tongjidata_NaN = pd.DataFrame([NaN_value.values], columns=NaN_value.index).to_dict('records')
    NaN_value = df.eq('X').sum()
    All_tongjidata_NaN = pd.DataFrame([NaN_value.values], columns=NaN_value.index).to_dict('records')
    # print(All_tongjidata_NaN)
    P_value_num = 0
    F_value_num = 0
    B_value_num = 0
    NS_value_num = 0
    Na_value_num = 0
    lienum = 0
    for i in key_data:
        if lienum >= 1:  # 因为读文件时时从第二列开始，要统计的时第三列开始的值
            if All_tongjidata_P[0][i] > 0:
                P_value_num += All_tongjidata_P[0][i] - 1
            F_value_num += All_tongjidata_F[0][i]
            B_value_num += All_tongjidata_B[0][i]
            NS_value_num += All_tongjidata_NS[0][i]
            Na_value_num += All_tongjidata_NaN[0][i]
        lienum += 1
    CaseStatus = ""
    # print(P_value_num, F_value_num, B_value_num, NS_value_num, Na_value_num)
    if F_value_num > 0:
        CaseStatus = "Fail"
    else:
        CaseStatus = "Pass"
    TestProess = (P_value_num + F_value_num) / (P_value_num + F_value_num + B_value_num + Na_value_num)
    TestProess = "%.2f%%" % (TestProess * 100)

    # 读取所有批注
    workbook = load_workbook(src_file)
    first_sheet = workbook.get_sheet_names()[1]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    comments = []
    rownum = 0
    for row in worksheet.rows:
        cellnum = 0
        for cell in row:
            if cell.comment:
                comments.append(["\n" + cell.comment.text])
            cellnum += 1
        rownum += 1
    # print(comments)

    # print("PPP", CaseStatus, TestProess)
    # print(datetime.datetime.now(),3)
    # return excel_dic, key_data
    file_name = src_file.split("/")[-1]
    # print(file_name)
    return excel_dic, key_data, CaseStatus, TestProess, comments, file_name

# def info_excel_tongji(src_file,header=0,sheetnum=1):
#     print(datetime.datetime.now(),2)
#     #您可以使用参数keep_default_na 和na_values 手动设置所有NA 值docs：防止pandas在读取excel时删除'NA‘字符串
#     df = pd.read_excel(src_file, header=header, sheet_name=int(sheetnum),keep_default_na=False).iloc[42:,
#          1:]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
#     # # 显示所有列
#     pd.set_option('display.max_columns', None)
#     # # 显示所有行
#     pd.set_option('display.max_rows', None)
#     # df = df.fillna('?')  # 替换 Nan, 否则没有双引号的Nan，json.dumps(data)时虽然不报错，但是传到前端反序列化后无法获取数据, None,NA,NAN,NAT,Null都被认为是缺失值
#     # df = df.fillna(method='ffill')
#     key_data = list(df.columns)
#     df = df.replace("", "?")
#     # pprint.pprint(df)
#
#     P_value = df.eq('P').sum()
#     All_tongjidata_P = pd.DataFrame([P_value.values], columns=P_value.index).to_dict('records')
#     F_value = df.eq('F').sum()
#     All_tongjidata_F = pd.DataFrame([F_value.values], columns=F_value.index).to_dict('records')
#     B_value = df.eq('B').sum()
#     All_tongjidata_B = pd.DataFrame([B_value.values], columns=B_value.index).to_dict('records')
#     NS_value = df.eq('NS').sum()
#     All_tongjidata_NS = pd.DataFrame([NS_value.values], columns=NS_value.index).to_dict('records')
#     # NaN_value = df.eq('?').sum()
#     # All_tongjidata_NaN = pd.DataFrame([NaN_value.values], columns=NaN_value.index).to_dict('records')
#     NaN_value = df.eq('X').sum()
#     All_tongjidata_NaN = pd.DataFrame([NaN_value.values], columns=NaN_value.index).to_dict('records')
#     # print(All_tongjidata_NaN)
#     P_value_num = 0
#     F_value_num = 0
#     B_value_num = 0
#     NS_value_num = 0
#     Na_value_num = 0
#     lienum = 0
#     for i in key_data:
#         if lienum >=1 :#因为读文件时时从第二列开始，要统计的时第三列开始的值
#             if All_tongjidata_P[0][i] > 0:
#                 P_value_num += All_tongjidata_P[0][i] - 1
#             F_value_num += All_tongjidata_F[0][i]
#             B_value_num += All_tongjidata_B[0][i]
#             NS_value_num += All_tongjidata_NS[0][i]
#             Na_value_num += All_tongjidata_NaN[0][i]
#         lienum += 1
#     CaseStatus = ""
#     # print(P_value_num, F_value_num, B_value_num, NS_value_num, Na_value_num)
#     if F_value_num > 0:
#         CaseStatus = "Fail"
#     else:
#         CaseStatus = "Pass"
#     TestProess = (P_value_num + F_value_num) / (P_value_num + F_value_num + B_value_num + Na_value_num)
#     TestProess = "%.2f%%" % (TestProess * 100)
#
#     # 读取所有批注
#     workbook = load_workbook(src_file)
#     first_sheet = workbook.get_sheet_names()[1]
#     worksheet = workbook.get_sheet_by_name(first_sheet)
#
#     comments = []
#     rownum = 0
#     for row in worksheet.rows:
#         cellnum = 0
#         for cell in row:
#             if cell.comment:
#                 comments.append(["\n", cell.comment.text])
#             cellnum += 1
#         rownum += 1
#     # print(comments)
#
#     # print("PPP", CaseStatus, TestProess)
#     print(datetime.datetime.now(),3)
#     return CaseStatus, TestProess, comments


def style_color(df, colors):
    """

    :param df: pd.DataFrame
    :param colors: 字典 内容是 {标题:颜色}
    :return:
    """
    return df.style.apply(style_apply, colors=colors)

def style_apply(series, colors, back_ground=''):
    """
    :param series: 传过来的数据是DataFramt中的一列  类型为pd.Series
    :param colors: 内容是字典 其中key 为标题名  value 为颜色
    :param back_ground: 北京颜色
    :return:
    """
    series_name = series.name[0]
    a = list()
    # 为了给每一个单元格上色
    # for col in series:
    #     # 其中 col 为pd.DataFrame 中的 一个小单元格  大家可以根据不同需求为单元格设置不同的颜色
    #     # 获取什么一级标题获取什么颜色
    #     if series_name in colors:
    #         for title_name in colors:
    #             if title_name == series_name:
    #                 back_ground = 'background-color: ' + colors[title_name]
    #                 # '; border-left-color: #080808'
    #     a.append(back_ground)
    for col in series:
        # 其中 col 为pd.DataFrame 中的 一个小单元格  大家可以根据不同需求为单元格设置不同的颜色
        if col in colors.keys():
            back_ground = 'background-color: ' + colors[col]
        else:
            back_ground = 'background-color: #FFFFFF'
                # '; border-left-color: #080808'
        a.append(back_ground)
    return a

def save_exel(folder_path_Sys,save_data,upload_zhushi,upload_zhushi_delete,src_file,auther,header=0,sheetnum=1):
    print("save_exel")
    # 读取所有批注
    workbook = load_workbook(src_file, data_only=False)
    first_sheet = workbook.get_sheet_names()[1]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    #修改前，先把原来文件里面的comments记录下来
    # comments = []
    excel_fx = []
    rownum = 0
    for row in worksheet.rows:
        cellnum = 0
        for cell in row:
            cell_value = worksheet.cell(row=rownum + 1, column=cellnum + 1).value
            if cell_value:
                if str(cell_value).startswith('='):
                    # print(cell_value)
                    excel_fx.append([rownum+1, cellnum+1, cell_value])
            # if cell.comment:
            #     # print(row, cell)
            #     comments.append([rownum, cellnum, cell.comment.text])
            cellnum += 1
        rownum += 1
    # print(comments)

    # 设置线条的样式和颜色
    side = Side(
                style="medium",  # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
                color="000000")
    # 设置单元格的边框线条
    border = Border(top=side, bottom=side, left=side, right=side)
    # 设置宽高
    # # row_dimensions中指定要设置高度的行
    # ws.row_dimensions[2].height = 50
    # column_dimensions中指定要设置宽度的列
    worksheet.column_dimensions['B'].width = 80

    fill_P = PatternFill(
        patternType="solid",  # 填充类型，可选none、solid、darkGray、mediumGray、lightGray、lightDown、lightGray、lightGrid
        fgColor="0000FF00",  # 前景色，16进制rgb
        bgColor="0000FF00",  # 背景色，16进制rgb
        # fill_type=None,     # 填充类型
        # start_color=None,   # 前景色，16进制rgb
        # end_color=None      # 背景色，16进制rgb
    )
    fill_F = PatternFill(
        patternType="solid",  # 填充类型，可选none、solid、darkGray、mediumGray、lightGray、lightDown、lightGray、lightGrid
        fgColor="00FF0000",  # 前景色，16进制rgb
        bgColor="00FF0000",  # 背景色，16进制rgb
    )
    fill_B = PatternFill(
        patternType="solid",  # 填充类型，可选none、solid、darkGray、mediumGray、lightGray、lightDown、lightGray、lightGrid
        fgColor="00FFFFCC",  # 前景色，16进制rgb
        bgColor="00FFFFCC",  # 背景色，16进制rgb
    )
    fill_NS = PatternFill(
        patternType="solid",  # 填充类型，可选none、solid、darkGray、mediumGray、lightGray、lightDown、lightGray、lightGrid
        fgColor="00C0C0C0",  # 前景色，16进制rgb
        bgColor="00C0C0C0",  # 背景色，16进制rgb
    )
    fill_Normal = PatternFill(
        patternType="solid",  # 填充类型，可选none、solid、darkGray、mediumGray、lightGray、lightDown、lightGray、lightGrid
        fgColor="00FFFFFF",  # 前景色，16进制rgb
        bgColor="00FFFFFF",  # 背景色，16进制rgb
    )
    print(save_data)
    for i in save_data:
        worksheet.cell(i["row"] + 2, i["lienum"] + 1).value = i["value"]
    # ws['A1'].alignment = Alignment(wrap_text=True)
    rownum = 0
    for row in worksheet:
        cellnum = 0
        for cell in row:
            # print(cell.value)
            if cell.value == "P":
                cell.fill = fill_P
            elif cell.value == "B":
                cell.fill = fill_B
            elif cell.value == "F":
                cell.fill = fill_F
            elif cell.value == "NS":
                cell.fill = fill_NS
            else:
                cell.fill = fill_Normal
            cell.border = border
            # 设置单元格自动换行
            cell.alignment = Alignment(wrap_text=True)
            cellnum += 1
        rownum += 1

    comments = upload_zhushi + upload_zhushi_delete
    for i in comments:
        hangnum = str(i[0] + 1)
        lie_num = str(ExcelColumn(i[1] + 1))
        comment = Comment(str(i[2]), str(auther))
        if i[2]:
            worksheet[lie_num + hangnum].comment = comment
        else:
            worksheet[lie_num + hangnum].comment = None
    # 保存公式最好放在保存前最后一步
    # print(excel_fx,'fx')
    for i in excel_fx:
        # print(i)
        worksheet.cell(row=i[0], column=i[1], value=i[2])
    workbook.save(src_file)
    # print(src_file, folder_path_Sys)
    shutil.copy(src_file, folder_path_Sys)

# def recursion_dir_all_file(path):
#     '''
#     :param path: 文件夹目录
#     '''
#     file_list = []
#     for dir_path, dirs, files in os.walk(path):
#         for file in files:
#             file_path = os.path.join(dir_path, file)
#             if "\\" in file_path:
#                 file_path = file_path.replace('\\', '/')
#             file_list.append(file_path)
#         for dir in dirs:
#             file_list.extend(recursion_dir_all_file(os.path.join(dir_path, dir)))
#     return file_list

@csrf_exempt
def ABOTestPlan_search(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi="ABOTestPlan/edit"
    # status='0'
    filepath = request.session.get('sessionABOEdit')
    # print(filepath,222)

    excel_dic = []
    key_list = []
    comments = []
    showinfo = ''
    canExport = 1
    canEdit = 0
    # TestID = id

    # print(request.method)
    if request.method == "POST":

        # print(request.POST, request.method)
        # print(request.body)
        # print(type(request.body),type(request.POST))
        # responseData = json.loads(request.body)
        # print(responseData)
        # if request.POST.get('isGetData') == 'first':

        # test = request.POST
        # for i in test:
        #     print(test[i])
        if request.POST.get('isGetData') == 'first':
            # folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
            if filepath:
                if os.path.exists(filepath):
                    readdata = read_excel(filepath)
                    excel_dic = readdata[0]
                    key_list = readdata[1]
                    comments = readdata[2]
                    showinfo = filepath.replace(settings.MEDIA_ROOT.replace('\\','/') + '/ABOTestPlanSys/', "")

        data = {
            "err_ok": "0",
            "excel_dic": excel_dic,
            "key_list": key_list,
            "comments": comments,
            "showinfo": showinfo,
            "canExport": canExport,
            "canEdit": canEdit,
            # "status":status
        }
        # print(type(json.dumps(data)),json.dumps(data))
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'ABOTestPlan/ABOTestPlan_search.html', locals())

@csrf_exempt
def ABOTestPlan_edit(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi="ABOTestPlan/edit"
    # status='0'
    filepath = request.session.get('sessionABOEdit')
    # print(filepath,222)

    excel_dic = []
    key_list = []
    canExport = 1
    canEdit = 0
    showinfo = ''
    errMsg = ''

    selectItem = {
        # "C38(NB)": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "C38(AIO)": [{"Project": "EL535", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL536", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL537", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL538", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "A39": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "Other": [{"Project": "ELMV2", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV3", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV4", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}]
    }
    selectcase = []
    tabledata = [
        # {"TestID": "ME-ENV001", "TestItems": "Operation Temperature Test1", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Pass", "Percent": "100%", "BugNo": "", "filepath": "/ABOTestPlan/ABOTestPlan_edit/",},
        # {"TestID": "ME-ENV002", "TestItems": "Operation Temperature Test2", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Fail", "Percent": "90%", "BugNo": "bug-212096.panel白斑", "filepath": "",},
        # {"TestID": "ME-ENV003", "TestItems": "Operation Temperature Test3", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Block", "Percent": "0%", "BugNo": "", "filepath": "",},
        # {"TestID": "ME-ENV004", "TestItems": "Operation Temperature Test4", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "NS", "Percent": "", "BugNo": "tttttttttttttttttttt", "filepath": "",},
    ]
    excel_dic = []
    key_list = []
    canExport = 1
    canEdit = 0
    err_msg = ""

    folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
    folder_path_Sys = settings.MEDIA_ROOT + '/ABOTestPlanSys/'  # 指定文件夹路径
    subforders = []
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for dirname in dirnames:
            subforders.append(os.path.join(dirpath, dirname).replace(folder_path, "").replace("\\", "/"))
            # print(os.path.join(dirpath, dirname).replace(folder_path, ""))
    subforders_name1 = []
    subforders_name2 = []
    subforders_name3 = []
    for i in subforders:
        if len(i.split("/")) == 1:
            subforders_name1.append(i)
        elif len(i.split("/")) == 2:
            subforders_name2.append(i)
        elif len(i.split("/")) == 3:
            subforders_name3.append(i)
        else:
            # print(i)
            pass
    for i in subforders_name1:
        projectincustomer = []
        for j in subforders_name2:
            phaseinproject = []
            if i in j:
                for k in subforders_name3:
                    if j in k:
                        phaseinproject.append(k.replace(j + "/", ""))
                projectincustomer.append({"Project": j.replace(i + "/", ""), "Phase0": phaseinproject})
        selectItem[i] = projectincustomer

    onlineuser = request.session.get('account')
    roles = []
    if UserInfo.objects.filter(account=onlineuser).first():
        for i in UserInfo.objects.filter(account=onlineuser).first().role.all():
            roles.append(i.name)
    # print(roles)
    # editPpriority = 100
    for i in roles:
        if 'admin' == i or 'DQA_ABO_Admin' == i or 'DQA_ABO_PL' == i or 'DQA_ABO_Tester' == i:
            canEdit = 1

    folder_path_Sys = settings.MEDIA_ROOT + '/ABOTestPlanSys/'  # 指定文件夹路径
    # TestID = id

    if request.method == "POST":

        # print(request.POST, request.method)
        # print(request.body)
        # print(type(request.body),type(request.POST))
        # responseData = json.loads(request.body)
        # print(responseData)
        # if request.POST.get('isGetData') == 'first':

        # test = request.POST
        # for i in test:
        #     print(test[i])
        if request.POST:
            if 'first' in str(request.body):
                try:
                    # folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
                    # # print(Customer, Category,111)
                    # filepathsearch = folder_path + Customer + "/" + Project + "/" + Phase + "/" + Category + "/" + cases
                    # filepathsearch = filepathsearch.replace("\\", "/").replace("//", "/")
                    # folder_path_Sys = folder_path_Sys + "%s_%s_%s_%s" % (
                    # Customer, Project, Phase, Category) + "/" + cases
                    # folder_path_Sys = folder_path_Sys.replace("\\", "/").replace("//", "/")
                    # # folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
                    # shutil.copy(filepathsearch, folder_path_Sys)
                    # folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
                    if filepath:
                        if os.path.exists(filepath):
                            readdata = read_excel(filepath)
                            excel_dic = readdata[0]
                            key_list = readdata[1]
                            comments = readdata[2]
                            showinfo = filepath.replace(settings.MEDIA_ROOT.replace('\\', '/') + '/ABOTestPlanSys/', "")

                except Exception as e:
                    print(e)
                    errMsg = str(e)
            if request.POST.get('isGetData') == 'changeCategorys':
                Customer = request.POST.get('Customer')
                Project = request.POST.get('Project')
                Phase = request.POST.get('Phase')
                Category = request.POST.get('Categorys')
                # print(Customer, Category,111)
                folder_path = folder_path + Customer + "/" + Project + "/" + Phase + "/" + Category
                folder_path = folder_path.replace("\\", "/").replace("//", "/")
                # folder_path_Sys = folder_path_Sys + "%s_%s_%s_%s" % (Customer, Project, Phase, Category)
                # folder_path_Sys = folder_path_Sys.replace("\\", "/").replace("//", "/")
                try:
                    for path in os.listdir(folder_path):
                        selectcase.append(path)
                except Exception as e:
                    print(e)
                    err_msg = str(e)
            if 'SEARCH' in str(request.body):
                try:
                    Customer = request.POST.get('Customer')
                    Project = request.POST.get('Project')
                    Phase = request.POST.get('Phase')
                    Category = request.POST.get('Categorys')
                    cases = request.POST.get('cases')
                    folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
                    # print(Customer, Category,111)
                    filepathsearch = folder_path + Customer + "/" + Project + "/" + Phase + "/" + Category + "/" + cases
                    filepathsearch = filepathsearch.replace("\\", "/").replace("//", "/")
                    folder_path_Sys = folder_path_Sys + "%s_%s_%s_%s" % (Customer, Project, Phase, Category) + "/" + cases
                    folder_path_Sys = folder_path_Sys.replace("\\", "/").replace("//", "/")
                    # folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
                    shutil.copy(filepathsearch, folder_path_Sys)
                    # print(folder_path_Sys)
                    if folder_path_Sys:
                        if os.path.exists(folder_path_Sys):
                            readdata = read_excel(folder_path_Sys)
                            excel_dic = readdata[0]
                            key_list = readdata[1]
                            comments = readdata[2]
                            showinfo = folder_path_Sys.replace(settings.MEDIA_ROOT.replace('\\', '/') + '/ABOTestPlanSys/', "")
                            request.session['sessionABOEdit'] = folder_path_Sys
                            request.session.set_expiry(12 * 60 * 60)
                            # print(showinfo,1)
                except Exception as e:
                    print(e)
                    errMsg = str(e)
        else:
            try:
                request.body
                # print(request.body)
            except:
                # print('1')
                pass
            else:
                if 'save' in str(request.body):
                    try:
                        responseData = json.loads(request.body)
                        uploaddata = responseData['uploadData']
                        upload_zhushi = responseData['upload_zhushi']
                        upload_zhushi_delete = responseData['upload_zhushi_delete']
                        showinfo = responseData['showinfo']
                        folder_path_Sys = filepath
                        # print(uploaddata, 'uploaddata')
                        # print(upload_zhushi, 'upload_zhushi')
                        # print(upload_zhushi_delete, 'upload_zhushi_delete')
                        # print(folder_path_Sys, 'folder_path_Sys')
                        oldstr = filepath.split("/")[-1]
                        folder_path_Sys_summary = filepath.replace(oldstr, '1Summary.xlsx')
                        strs = filepath.split("ABOTestPlanSys")
                        # print(strs)
                        val = strs[1].count('_')
                        # print(strs[1].replace("_", "/", val-1))
                        filepath = strs[0] + "ABOTestPlan" + strs[1].replace("_", "/", val - 1)
                        # print(filepath,os.path.exists(filepath))
                        # filepath = filepath.replace("_", "/", val-1).replace('/ABOTestPlanSys/', '/ABOTestPlan/')
                        # print(filepath)
                        auther = request.session.get('user_name')
                        if filepath:
                            if os.path.exists(filepath):
                                # print('start')
                                save_exel(folder_path_Sys,uploaddata,upload_zhushi,upload_zhushi_delete, filepath,auther)
                                readdata = read_excel(filepath)
                                excel_dic = readdata[0]
                                key_list = readdata[1]
                                comments = readdata[2]
                        #每次保存时将数据更新到1Summry.xlsx中去，但是当里面有几百条数据时好像也挺花时间的，影响save的时间
                        # print(folder_path_Sys_summary)
                        if os.path.exists(folder_path_Sys):
                            # print(oldstr,folder_path_Sys_summary)
                            excel_data = []
                            excel_info = info_excel(folder_path_Sys)
                            # print(excel_info)
                            # 将列表转换为DataFrame
                            excel_data.append(excel_info)
                            tabledata1 = []
                            for i in excel_data:
                                # print(i)
                                key = i[5]
                                SKU = i[0][0]['SKU/Unit']
                                Owner = i[0][0]['Owner']
                                TestSchedule = i[0][0]['Test Schedule']
                                # print(excel_dic)
                                tabledata1.append(
                                    {
                                        "TestID": key.split(".")[0].split("_")[0],
                                        "TestItems": key.split(".")[0].split("_")[1],
                                        "SKU": SKU, "Owner": Owner, "TestSchedule": TestSchedule,
                                        # "Status": Result[0], "Percent": Result[1], "BugNo": Result[2],
                                        "Status": i[2], "Percent": i[3], "BugNo": i[4],
                                        "filepath": folder_path_Sys,
                                    }
                                )
                        # print(tabledata1)
                        if not os.path.exists(folder_path_Sys_summary):
                            # 将列表转换为DataFrame
                            df = pd.DataFrame(tabledata1)
                            # 将DataFrame输出到Excel文件
                            df.to_excel(folder_path_Sys_summary, index=False)
                        else:
                            df = pd.read_excel(folder_path_Sys_summary, header=0, sheet_name=0).iloc[:,
                                 :]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
                            rownum = df[(df.TestItems == tabledata1[0]["TestItems"])].index.tolist()
                            allrownum = df.shape[0] - 1
                            # print(filepath)
                            # print(tabledata1[0])
                            # print(rownum, allrownum)
                            if rownum:
                                # print("tihuan")
                                df = df.drop(index=rownum)
                                # print(df,"drop")
                                df_part1 = df.loc[0:rownum[0] - 1]
                                df_part2 = df.loc[rownum[0]:allrownum]
                                dffinal = df_part1.append(tabledata1[0], ignore_index=True)
                                dffinal = dffinal.append(df_part2, ignore_index=True)
                            else:
                                # print(df,1)
                                # print(pd.DataFrame(tabledata))
                                dffinal = df.append(tabledata1[0], ignore_index=True)
                                # print(pd.DataFrame(tabledata1[0]))
                                # print(df)

                            dffinal.sort_values(by=["TestID"], inplace=True)  # 多条件从小到大
                            dffinal.reset_index(drop=True)
                            # print(dffinal)
                            dffinal.to_excel(folder_path_Sys_summary, index=False)
                    except Exception as e:
                        print(e)
                        errMsg = str(e)

                    # save_exel(excel_dic,src_file)
        data = {
            "err_ok": "0",
            "errMsg": errMsg,
            "excel_dic": excel_dic,
            "key_list": key_list,
            "canExport": canExport,
            "canEdit": canEdit,
            "showinfo": showinfo,
            # "status":status
            "select": selectItem,
            "selectcase": selectcase,

        }
        # print(type(json.dumps(data)),json.dumps(data))
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'ABOTestPlan/ABOTestPlan_edit.html', locals())

@csrf_exempt
def ABOTestPlan_summary(request):
    if not request.session.get('is_login', None):
        return redirect('/login/')
    Skin = request.COOKIES.get('Skin_raw')
    # print(Skin)
    if not Skin:
        Skin = "/static/src/blue.jpg"
    weizhi="ABOTestPlan/Summary"
    # status='0'
    selectItem = {
        # "C38(NB)": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                       {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "C38(AIO)": [{"Project": "EL535", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL536", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL537", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                        {"Project": "EL538", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "A39": [{"Project": "EL531", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL532", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL533", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                   {"Project": "EL534", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}],
        #           "Other": [{"Project": "ELMV2", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV3", "Phase0": ["B(FVT)", "C(SIT)", "INV"]},
        #                     {"Project": "ELMV4", "Phase0": ["B(FVT)", "C(SIT)", "INV"]}]
    }
    tabledata = [
        # {"TestID": "ME-ENV001", "TestItems": "Operation Temperature Test1", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Pass", "Percent": "100%", "BugNo": "", "filepath": "/ABOTestPlan/ABOTestPlan_edit/",},
        # {"TestID": "ME-ENV002", "TestItems": "Operation Temperature Test2", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Fail", "Percent": "90%", "BugNo": "bug-212096.panel白斑", "filepath": "",},
        # {"TestID": "ME-ENV003", "TestItems": "Operation Temperature Test3", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "Block", "Percent": "0%", "BugNo": "", "filepath": "",},
        # {"TestID": "ME-ENV004", "TestItems": "Operation Temperature Test4", "SKU": "1a#14~16,2a#7~8", "Owner": "DQA", "Schedule": "9/8~9/14",
        #  "Status": "NS", "Percent": "", "BugNo": "tttttttttttttttttttt", "filepath": "",},
    ]
    excel_dic = []
    key_list = []
    canExport = 1
    canEdit = 0
    onlineuser = request.session.get('account')
    roles = []
    if UserInfo.objects.filter(account=onlineuser).first():
        for i in UserInfo.objects.filter(account=onlineuser).first().role.all():
            roles.append(i.name)
    # print(roles)
    # editPpriority = 100
    for i in roles:
        if 'admin' == i or 'DQA_ABO_Admin' == i or 'DQA_ABO_PL' == i or 'DQA_ABO_Tester' == i:
            canEdit = 1
    # canExport = canEdit
    err_msg = ""

    folder_path = settings.MEDIA_ROOT + '/ABOTestPlan/'  # 指定文件夹路径
    folder_path_Sys = settings.MEDIA_ROOT + '/ABOTestPlanSys/'  # 指定文件夹路径
    subforders = []
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for dirname in dirnames:
            subforders.append(os.path.join(dirpath, dirname).replace(folder_path, "").replace("\\", "/"))
            # print(os.path.join(dirpath, dirname).replace(folder_path, ""))
    subforders_name1 = []
    subforders_name2 = []
    subforders_name3 = []
    for i in subforders:
        if len(i.split("/")) == 1:
            subforders_name1.append(i)
        elif len(i.split("/")) == 2:
            subforders_name2.append(i)
        elif len(i.split("/")) == 3:
            subforders_name3.append(i)
        else:
            # print(i)
            pass
    for i in subforders_name1:
        projectincustomer = []
        for j in subforders_name2:
            phaseinproject = []
            if i in j:
                for k in subforders_name3:
                    if j in k:
                        phaseinproject.append(k.replace(j + "/", ""))
                projectincustomer.append({"Project": j.replace(i + "/", ""), "Phase0": phaseinproject})
        selectItem[i] = projectincustomer
    # print(selectItem)
    # print(request.method)
    if request.method == "POST":
        if request.POST.get('isGetData') == 'first':
            pass

        if request.POST.get('isGetData') == 'SEARCH':
            Customer = request.POST.get('Customer')
            Project = request.POST.get('Project')
            Phase = request.POST.get('Phase')
            Category = request.POST.get('Categorys')
            # print(Customer, Category,111)
            folder_path = folder_path + Customer + "/" + Project + "/" + Phase + "/" + Category
            folder_path = folder_path.replace("\\", "/").replace("//", "/")
            folder_path_Sys = folder_path_Sys + "%s_%s_%s_%s" % (Customer, Project, Phase, Category)
            folder_path_Sys = folder_path_Sys.replace("\\", "/").replace("//", "/")
            folder_path_Sys_summary = folder_path_Sys + "/" + '1Summary.xlsx'
            try:
                file_ext = ['.xls', '.xlsx']
                if not os.path.exists(folder_path_Sys_summary):
                    # print(datetime.datetime.now())
                    Copy_forders(folder_path, folder_path_Sys)
                    # print(datetime.datetime.now())
                    #
                    # 也可以使用 with 语句创建线程池
                    # with ThreadPoolExecutor(max_workers=3) as pool:
                    #     for i in range(1, 14):
                    #         pool.submit(async_add, i)
                    # pool = ThreadPoolExecutor(6)#多线程第一步：创建8个线程，由于python GIL锁，导致多线程速度更慢了
                    # 实例化一个线程池对象，max_workers 设置线程池中能同时运行的最大线程数目，如果不指定默认是 cpu 核数的5倍，thread_name_prefix用来指定线程名前缀
                    pool = ThreadPoolExecutor(max_workers=60, thread_name_prefix='Excel_info')#多线程第一步：创建8个线程，由于python GIL锁，导致多线程速度更慢了
                    # pool = ProcessPoolExecutor(6)# 实例化获得一个进程池, 参数传入一个整数，代表进程池的大小,不传的话会默认开设当前计算机CPU 个数的进程
                    # pool1 = ThreadPoolExecutor(6)#多线程第一步：创建8个线程，由于python GIL锁，导致多线程速度更慢了
                    # 实例化一个线程池对象，max_workers 设置线程池中能同时运行的最大线程数目，如果不指定默认是 cpu 核数的5倍，thread_name_prefix用来指定线程名前缀
                    # pool1 = ThreadPoolExecutor(max_workers=30, thread_name_prefix='Excel_tongji')#多线程第一步：创建8个线程，由于python GIL锁，导致多线程速度更慢了
                    # pool1 = ProcessPoolExecutor(6)  # 实例化获得一个进程池, 参数传入一个整数，代表进程池的大小,不传的话会默认开设当前计算机CPU 个数的进程
                    if os.path.exists(folder_path_Sys):
                        # i = 0
                        all_result = []
                        for path in os.listdir(folder_path_Sys):
                            if "1Summary" not in path:
                                path_list = os.path.join(folder_path_Sys, path)  # 连接当前目录及文件或文件夹名称
                                path_list = path_list.replace("\\", "/")
                                if os.path.isfile(path_list):  # 判断当前文件或文件夹是否是文件，把文件夹排除
                                    if (os.path.splitext(path_list)[1]) in file_ext:  # 判断取得文件的扩展名是否是.xls、.xlsx
                                        # print(path_list, path)  # 打印输出
                                        excel_dic_process = pool.submit(info_excel, path_list)#多线程第二步：在线程池中发任务，由于python GIL锁，导致多线程速度更慢了
                                        excel_dic = excel_dic_process.result()#多线程第二步：接收多线程执行函数的返回值，由于python GIL锁，导致多线程速度更慢了
                                        # Result_process = pool1.submit(info_excel_tongji, path_list)  # 多线程第二步：在线程池中发任务，由于python GIL锁，导致多线程速度更慢了
                                        # Result = Result_process.result()  # 多线程第二步：接收多线程执行函数的返回值，由于python GIL锁，导致多线程速度更慢了
                                        # Result = info_excel_tongji(path_list)
                                        # excel_dic = info_excel(path_list)
                                        # print(excel_dic)
                                        all_result.append(excel_dic)

                                        # i += 1  # 对.xls、.xlsx文件进行计数
                        # pool.shutdown()
                        pool.shutdown(wait=True)#多线程第三步：等待线程池把任务都执行完毕，由于python GIL锁，导致多线程速度更慢了
                        # pool1.shutdown(wait=True)#多线程第三步：等待线程池把任务都执行完毕，由于python GIL锁，导致多线程速度更慢了
                        # print('目录下共有' + str(i) + '个xls、xlsx文件')
                        #
                        # Del_forders(folder_path_Sys)
                    for i in all_result:
                        # print(i)
                        key = i[5]
                        SKU = i[0][0]['SKU/Unit']
                        Owner = i[0][0]['Owner']
                        TestSchedule = i[0][0]['Test Schedule']
                        # print(excel_dic)
                        path_list = os.path.join(folder_path_Sys, key)  # 连接当前目录及文件或文件夹名称
                        path_list = path_list.replace("\\", "/")
                        tabledata.append(
                            {
                                "TestID": key.split(".")[0].split("_")[0],
                                "TestItems": key.split(".")[0].split("_")[1],
                                "SKU": SKU, "Owner": Owner, "TestSchedule": TestSchedule,
                                # "Status": Result[0], "Percent": Result[1], "BugNo": Result[2],
                                "Status": i[2], "Percent": i[3], "BugNo": i[4],
                                "filepath": path_list,
                            }
                        )

                        # 将列表转换为DataFrame
                        df = pd.DataFrame(tabledata)
                        # 将DataFrame输出到Excel文件
                        df.to_excel(folder_path_Sys_summary, index=False)
                else:
                    Copy_forders(folder_path, folder_path_Sys)
                    df = pd.read_excel(folder_path_Sys_summary, header=0, sheet_name=0).iloc[:,
                         :]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
                    # # 显示所有列
                    pd.set_option('display.max_columns', None)
                    # # 显示所有行
                    # pd.set_option('display.max_rows', None)
                    # pprint.pprint(df)
                    dataexcel = df.values[:, :]
                    # datatest = []
                    # for i in dataexcel:
                    #     ls = []
                    #     for j in i:
                    #         ls.append(j)
                    #     datatest.append(ls)
                    # print(list(df.columns))
                    # pprint.pprint(datatest)
                    excel_dic1 = df.to_dict('records')
                    # print("111", excel_dic1)
                    # key_data = list(df.columns)
                    for i in excel_dic1:
                        # print(i)
                        tabledata.append(
                            {
                                "TestID": i["TestID"],
                                "TestItems": i["TestItems"],
                                "SKU": i["SKU"], "Owner": i["Owner"], "TestSchedule": i["TestSchedule"],
                                # "Status": Result[0], "Percent": Result[1], "BugNo": Result[2],
                                "Status": i["Status"], "Percent": i["Percent"], "BugNo": i["BugNo"].replace("[", "").replace("]", "").replace("'", "").replace("\\n", "\n"),
                                "filepath": i["filepath"],
                            }
                        )

                    pool = ThreadPoolExecutor(max_workers=60, thread_name_prefix='Excel_info')
                    if os.path.exists(folder_path_Sys):
                        # i = 0
                        filenames_notinsummary = []
                        for path in os.listdir(folder_path):
                            if path not in str(tabledata):
                                filenames_notinsummary.append(path)
                        # print(filenames_notinsummary)
                        all_result = []
                        if filenames_notinsummary:
                            for path in filenames_notinsummary:
                                    path_list = os.path.join(folder_path_Sys, path)  # 连接当前目录及文件或文件夹名称
                                    path_list = path_list.replace("\\", "/")
                                    if os.path.isfile(path_list):  # 判断当前文件或文件夹是否是文件，把文件夹排除
                                        if (os.path.splitext(path_list)[1]) in file_ext:  # 判断取得文件的扩展名是否是.xls、.xlsx
                                            # print(path_list, path)  # 打印输出
                                            excel_dic_process = pool.submit(info_excel, path_list)#多线程第二步：在线程池中发任务，由于python GIL锁，导致多线程速度更慢了
                                            excel_dic = excel_dic_process.result()#多线程第二步：接收多线程执行函数的返回值，由于python GIL锁，导致多线程速度更慢了
                                            # Result_process = pool1.submit(info_excel_tongji, path_list)  # 多线程第二步：在线程池中发任务，由于python GIL锁，导致多线程速度更慢了
                                            # Result = Result_process.result()  # 多线程第二步：接收多线程执行函数的返回值，由于python GIL锁，导致多线程速度更慢了
                                            # Result = info_excel_tongji(path_list)
                                            # excel_dic = info_excel(path_list)
                                            # print(excel_dic)
                                            all_result.append(excel_dic)

                                            # i += 1  # 对.xls、.xlsx文件进行计数
                            # pool.shutdown()
                            pool.shutdown(wait=True)#多线程第三步：等待线程池把任务都执行完毕，由于python GIL锁，导致多线程速度更慢了
                            # pool1.shutdown(wait=True)#多线程第三步：等待线程池把任务都执行完毕，由于python GIL锁，导致多线程速度更慢了
                            # print('目录下共有' + str(i) + '个xls、xlsx文件')
                            tabledata_notinsummary = []
                            if all_result:
                                for i in all_result:
                                    # print(i)
                                    key = i[5]
                                    SKU = i[0][0]['SKU/Unit']
                                    Owner = i[0][0]['Owner']
                                    TestSchedule = i[0][0]['Test Schedule']
                                    # print(excel_dic)
                                    path_list = os.path.join(folder_path_Sys, key)  # 连接当前目录及文件或文件夹名称
                                    path_list = path_list.replace("\\", "/")
                                    tabledata.append(
                                        {
                                            "TestID": key.split(".")[0].split("_")[0],
                                            "TestItems": key.split(".")[0].split("_")[1],
                                            "SKU": SKU, "Owner": Owner, "TestSchedule": TestSchedule,
                                            # "Status": Result[0], "Percent": Result[1], "BugNo": Result[2],
                                            "Status": i[2], "Percent": i[3], "BugNo": i[4],
                                            "filepath": path_list,
                                        }
                                    )
                                    tabledata_notinsummary.append(
                                        {
                                            "TestID": key.split(".")[0].split("_")[0],
                                            "TestItems": key.split(".")[0].split("_")[1],
                                            "SKU": SKU, "Owner": Owner, "TestSchedule": TestSchedule,
                                            # "Status": Result[0], "Percent": Result[1], "BugNo": Result[2],
                                            "Status": i[2], "Percent": i[3], "BugNo": i[4],
                                            "filepath": path_list,
                                        }
                                    )
                                if tabledata_notinsummary:
                                    df = pd.read_excel(folder_path_Sys_summary, header=0, sheet_name=0).iloc[:,
                                         :]  # ‘,’前面是行，后面是列，sheet_name指定sheet，可是是int第几个，可以是名称，header从第几行开始读取
                                    # print(df,"drop")
                                    df = df.append(pd.DataFrame(tabledata_notinsummary), ignore_index=True)
                                    # print(pd.DataFrame(tabledata_notinsummary))
                                    # 将DataFrame输出到Excel文件
                                    df.sort_values(by=["TestID"], inplace=True)  # 多条件从小到大
                                    df.reset_index(drop=True)
                                    # print(dffinal)
                                    df.to_excel(folder_path_Sys_summary, index=False)

            except Exception as e:
                print(e)
                err_msg = str(e)



        if request.POST.get("isGetData") == "ABOTestPlan_edit":
            #cookie
            # Redirect = redirect('/Lesson_search/')
            # Compatibilityv = request.POST.get('isGetData')
            # Redirect.set_cookie('cookieSWME', Compatibilityv, 3600 * 24 )
            # return Redirect#这里虽然返回了Redirect的路径，但是由于时axios传输，返回页面没有用，到那时必须要加，不然cookie设置不成功。
            filepath = request.POST.get('filepath')
            # print(filepath, "111")
            request.session['sessionABOEdit'] = filepath
            request.session.set_expiry(12 * 60 * 60)
            return render(request, 'ABOTestPlan/ABOTestPlan_Summary.html')


        data = {
            "err_msg": err_msg,
            "tabledata": tabledata,
            "select": selectItem,
            "excel_dic": excel_dic,
            "key_list": key_list,
            "canExport": canExport,
            "canEdit": canEdit,
            # "status":status
        }
        # print(data)
        return HttpResponse(json.dumps(data), content_type="application/json")
    return render(request, 'ABOTestPlan/ABOTestPlan_Summary.html', locals())